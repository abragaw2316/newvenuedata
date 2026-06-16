"""Build a polished, agent-facing .xlsx from the generated lead-list JSON.

Standalone (NOT part of the zero-dep Node pipeline). Run after `npm run leads`:
    python validation/make-xlsx.py
Needs openpyxl (`pip install openpyxl`). Output: south-fl-new-liquor-leads.xlsx
"""
import json
from collections import Counter
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = Path(__file__).parent
leads = json.loads((HERE / "south-fl-new-liquor-leads.json").read_text(encoding="utf-8"))

# Brand palette (matches the New Venue Data indigo accent)
INDIGO = "4F46E5"
INDIGO_DK = "3730A3"
ROW_ALT = "EEF0FB"
GREY = "6B7280"
ARIAL = "Arial"

def as_date(s):
    try:
        return datetime.strptime(s, "%Y-%m-%d")
    except (ValueError, TypeError):
        return s or ""

COLS = [
    ("Business Name", "businessName", 30),
    ("License Type", "licenseType", 32),
    ("On-Prem", "onPremises", 9),
    ("License #", "licenseNumber", 13),
    ("Street", "street", 30),
    ("City", "city", 18),
    ("County", "county", 13),
    ("ZIP", "zip", 8),
    ("Filed", "filedDate", 12),
    ("Expires", "licenseExpires", 12),
    ("Licensee (Legal Name)", "licensee", 30),
]
DATE_KEYS = {"filedDate", "licenseExpires"}

wb = Workbook()
ws = wb.active
ws.title = "New Liquor Leads"
ws.sheet_view.showGridLines = False

filed = sorted(l["filedDate"] for l in leads if l.get("filedDate"))
span = f"{filed[0]} to {filed[-1]}" if filed else ""

# ── banner ───────────────────────────────────────────────────────────────────
last_col = len(COLS)
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
t = ws.cell(1, 1, "New Venue Data — New Liquor License Filings, South Florida")
t.font = Font(name=ARIAL, size=15, bold=True, color="FFFFFF")
t.fill = PatternFill("solid", fgColor=INDIGO)
t.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws.row_dimensions[1].height = 30

ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
s = ws.cell(2, 1, f"Sample lead list · {len(leads)} brand-new venues licensed to serve alcohol "
                  f"(Broward / Miami-Dade / Palm Beach) · filed {span}")
s.font = Font(name=ARIAL, size=10, color="FFFFFF")
s.fill = PatternFill("solid", fgColor=INDIGO_DK)
s.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws.row_dimensions[2].height = 20

ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=last_col)
d = ws.cell(3, 1, "Source: FL DBPR — Alcoholic Beverages & Tobacco (Chapter 119 public records). "
                  "Provided as-is; New Venue Data is not affiliated with or endorsed by DBPR or the State of Florida.")
d.font = Font(name=ARIAL, size=8, italic=True, color=GREY)
d.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws.row_dimensions[3].height = 16

HEADER_ROW = 5
thin = Side(style="thin", color="D1D5DB")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# ── header ───────────────────────────────────────────────────────────────────
for c, (title, _key, width) in enumerate(COLS, start=1):
    cell = ws.cell(HEADER_ROW, c, title)
    cell.font = Font(name=ARIAL, size=10, bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=INDIGO)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border
    ws.column_dimensions[get_column_letter(c)].width = width
ws.row_dimensions[HEADER_ROW].height = 26

# ── data ─────────────────────────────────────────────────────────────────────
for i, lead in enumerate(leads):
    r = HEADER_ROW + 1 + i
    shade = ROW_ALT if i % 2 else "FFFFFF"
    for c, (_title, key, _w) in enumerate(COLS, start=1):
        val = as_date(lead.get(key, "")) if key in DATE_KEYS else lead.get(key, "")
        cell = ws.cell(r, c, val)
        cell.font = Font(name=ARIAL, size=10)
        cell.fill = PatternFill("solid", fgColor=shade)
        cell.border = border
        cell.alignment = Alignment(
            horizontal="center" if key in {"onPremises", "zip", "licenseNumber"} | DATE_KEYS else "left",
            vertical="center",
        )
        if key in DATE_KEYS and isinstance(val, datetime):
            cell.number_format = "yyyy-mm-dd"

last_data_row = HEADER_ROW + len(leads)
ws.freeze_panes = f"A{HEADER_ROW + 1}"
ws.auto_filter.ref = f"A{HEADER_ROW}:{get_column_letter(last_col)}{last_data_row}"

# ── Summary sheet ─────────────────────────────────────────────────────────────
# Counts are computed here (a static snapshot of these sample rows; the sheet is
# regenerated wholesale by `npm run leads` + this script, not hand-edited in Excel,
# so a stored value stays correct and renders without a formula recalc step).
sm = wb.create_sheet("Summary")
sm.sheet_view.showGridLines = False
sm.column_dimensions["A"].width = 40
sm.column_dimensions["B"].width = 10

by_county = Counter(l["county"] for l in leads)
by_type = Counter(l["licenseType"] for l in leads)

def head(r, text):
    cell = sm.cell(r, 1, text)
    cell.font = Font(name=ARIAL, size=11, bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=INDIGO)
    sm.cell(r, 2).fill = PatternFill("solid", fgColor=INDIGO)

def line(r, label, value, bold=False):
    a = sm.cell(r, 1, label)
    a.font = Font(name=ARIAL, size=10 + bold, bold=bold)
    a.alignment = Alignment(wrap_text=True, vertical="center")
    b = sm.cell(r, 2, value)
    b.font = Font(name=ARIAL, size=10 + bold, bold=bold)
    b.alignment = Alignment(horizontal="center")

row = 1
line(row, "Total leads", len(leads), bold=True); row += 2
head(row, "Leads by county"); row += 1
for county, n in sorted(by_county.items()):
    line(row, county, n); row += 1
row += 1
head(row, "Leads by license type"); row += 1
for lt, n in sorted(by_type.items()):
    line(row, lt, n); row += 1

out = HERE / "south-fl-new-liquor-leads.xlsx"
wb.save(out)
print(f"wrote {out} ({len(leads)} leads, 2 sheets)")
